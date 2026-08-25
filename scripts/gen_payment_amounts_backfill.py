"""Generate SQL to backfill payment_amounts from Excel Deduction Dates."""
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX = Path(r"g:\Downloads\All_Loans_Readable_Summary_2026-08-21.xlsx")
OUT = Path("supabase/backfill_payment_amounts_from_excel.sql")

wb = openpyxl.load_workbook(XLSX, data_only=True)

# Deduction Dates: actual per-payment amounts
ws = wb["Deduction Dates"]
by_ref: dict[str, list[tuple[int, str, float]]] = defaultdict(list)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]:
        continue
    ref = str(r[1]).strip()
    ded_num = r[4]
    d = r[5]
    amt = r[6]
    if d is None or amt is None:
        continue
    if isinstance(d, datetime):
        ds = d.date().isoformat()
    elif isinstance(d, date):
        ds = d.isoformat()
    else:
        ds = str(d)[:10]
    by_ref[ref].append((int(ded_num) if ded_num else 0, ds, float(amt)))

# All Loans: closed / remaining
ws2 = wb["All Loans"]
status: dict[str, dict] = {}
for r in ws2.iter_rows(min_row=4, values_only=True):
    if not r or not r[1]:
        continue
    ref = str(r[1]).strip()
    status[ref] = {
        "remaining": float(r[8] or 0),
        "status": r[11] or "",
    }

lines = [
    "-- Backfill payment_amounts from Deduction Dates (uneven schedules).",
    "-- Closed loans: total_installments = actual deduction count.",
    "BEGIN;",
]

for ref, rows in sorted(by_ref.items()):
    rows = sorted(rows, key=lambda x: (x[0], x[1]))
    dates = [d for _, d, _ in rows]
    amounts = [round(a, 2) for _, _, a in rows]
    n = len(amounts)
    st = status.get(ref, {})
    closed = st.get("status") == "Closed" or float(st.get("remaining") or 0) == 0
    dates_json = json.dumps(dates)
    amounts_json = json.dumps(amounts)
    ref_sql = ref.replace("'", "''")

    if closed:
        lines.append(
            f"""
UPDATE public.loans SET
  payment_dates = '{dates_json}'::jsonb,
  payment_amounts = '{amounts_json}'::jsonb,
  paid_count = {n},
  total_installments = {n}
WHERE ref = '{ref_sql}';"""
        )
    else:
        lines.append(
            f"""
UPDATE public.loans SET
  payment_dates = '{dates_json}'::jsonb,
  payment_amounts = '{amounts_json}'::jsonb,
  paid_count = GREATEST(paid_count, {n})
WHERE ref = '{ref_sql}';"""
        )

lines.append("COMMIT;")
OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"wrote {OUT} ({len(by_ref)} refs)")

# Preview L315
for block in lines:
    if "L315" in block:
        print(block)
        break

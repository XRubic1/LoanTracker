"""Re-import OPEN Lexie/Neno loans from Excel. Never touch Team Petar."""
from __future__ import annotations

import json
import re
import subprocess
from datetime import date, datetime
from pathlib import Path

import openpyxl

EXCEL = Path(r"g:\Downloads\All_Loans_Readable_Summary_2026-08-21.xlsx")
REPO = Path(r"C:\Users\petar\OneDrive\Desktop\loantrucker")
OUT_SQL = REPO / "supabase" / "reimport_open_loans_lexi_neno_2026_08_21.sql"


def db_query(sql: str) -> list[dict]:
    proc = subprocess.run(
        ["npx", "supabase", "db", "query", "--linked", "-o", "json", sql],
        cwd=str(REPO),
        capture_output=True,
        text=True,
        shell=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr or proc.stdout)
    start = proc.stdout.find("{")
    if start < 0:
        raise RuntimeError(proc.stdout)
    return json.loads(proc.stdout[start:]).get("rows") or []


def q(value: object | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def norm_ref(ref: object) -> str:
    return re.sub(r"\s+", "", str(ref or "").strip().upper())


def to_iso_date(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None


def parse_paid(paid: object) -> tuple[int, int]:
    m = re.match(r"^(\d+)\s*/\s*(\d+)$", str(paid or "").strip())
    if not m:
        return 0, 1
    return int(m.group(1)), int(m.group(2))


def parse_deduction_dates(raw: object) -> list[str]:
    if raw is None or raw == "":
        return []
    if isinstance(raw, datetime):
        return [raw.date().isoformat()]
    out: list[str] = []
    for part in re.split(r"[;|]", str(raw)):
        part = part.strip()
        if not part:
            continue
        try:
            out.append(datetime.strptime(part, "%b %d, %Y").date().isoformat())
        except ValueError:
            d = to_iso_date(part)
            if d:
                out.append(d)
    return out


def load_rows(ws) -> list[dict]:
    headers = [ws.cell(3, c).value for c in range(1, 16)]
    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 16)]
        if vals[1]:
            rows.append(dict(zip(headers, vals)))
    return rows


def main() -> None:
    companies = {
        r["name"]: r["owner_id"]
        for r in db_query(
            "SELECT name, owner_id::text AS owner_id FROM public.companies ORDER BY name"
        )
    }
    print("Companies:", companies)

    owner_lexi = companies["TRUFUNDING Team Lexi"]
    owner_neno = companies["TRUFUNDING Team Nenad"]
    owner_petar = companies["TRUFUNDING Team Petar"]
    source_owner = {"Lexie": owner_lexi, "Neno": owner_neno}

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    all_rows = load_rows(wb["All Loans"])
    open_rows = load_rows(wb["Open Loans"])

    closed: dict[str, set[str]] = {owner_lexi: set(), owner_neno: set()}
    for row in all_rows:
        if str(row.get("Status") or "").strip().lower() != "closed":
            continue
        owner = source_owner.get(str(row.get("Source") or "").strip())
        if not owner:
            continue
        ref = norm_ref(row.get("Ref"))
        if ref:
            closed[owner].add(ref)

    open_import: list[tuple[str, dict]] = []
    seen: set[tuple[str, str]] = set()
    for row in open_rows:
        owner = source_owner.get(str(row.get("Source") or "").strip())
        if not owner:
            continue
        ref = norm_ref(row.get("Ref"))
        if not ref or (owner, ref) in seen:
            continue
        seen.add((owner, ref))
        open_import.append((owner, row))

    lines: list[str] = [
        "-- Re-import OPEN Lexie/Neno loans only; never touch Petar",
        f"-- Generated {datetime.now().isoformat(timespec='seconds')}",
        "BEGIN;",
        "",
    ]

    for owner, refs in closed.items():
        if not refs:
            continue
        ref_list = ", ".join(q(r) for r in sorted(refs))
        lines += [
            "DELETE FROM public.loans",
            f"WHERE owner_id = {q(owner)}::uuid",
            f"  AND upper(regexp_replace(coalesce(ref,''), '\\s+', '', 'g')) IN ({ref_list});",
            "",
        ]

    open_refs: dict[str, set[str]] = {owner_lexi: set(), owner_neno: set()}
    for owner, row in open_import:
        open_refs[owner].add(norm_ref(row.get("Ref")))
    for owner, refs in open_refs.items():
        if not refs:
            continue
        ref_list = ", ".join(q(r) for r in sorted(refs))
        lines += [
            "DELETE FROM public.loans",
            f"WHERE owner_id = {q(owner)}::uuid",
            f"  AND upper(regexp_replace(coalesce(ref,''), '\\s+', '', 'g')) IN ({ref_list});",
            "",
        ]

    lines.append(
        "INSERT INTO public.loans (\n"
        "  owner_id, client, ref, total, installment, paid_count, total_installments,\n"
        "  start_date, freq_days, payment_dates, payment_notes, payment_amounts,\n"
        "  partial_paid_amount, note, provider_type, provider_name, factoring_fee, hidden\n"
        ") VALUES"
    )

    values: list[str] = []
    excel_remaining = 0.0
    for owner, row in open_import:
        paid_count, total_inst = parse_paid(row.get("Paid"))
        total = float(row.get("Total") or 0)
        installment = float(row.get("Installment") or 0)
        excel_remaining += float(row.get("Remaining") or 0)
        start = to_iso_date(row.get("Start")) or date.today().isoformat()
        dates = parse_deduction_dates(row.get("Deduction dates"))
        if len(dates) > paid_count:
            dates = dates[:paid_count]
        while len(dates) < paid_count:
            dates.append(start)
        notes = [""] * max(total_inst, 1)

        provider_raw = str(row.get("Provider") or "").strip()
        if not provider_raw or provider_raw.lower() in ("trufunding", "tru funding"):
            provider_type, provider_name = "TruFunding", None
        else:
            provider_type, provider_name = "Other", provider_raw

        factoring = float(row.get("Factoring income") or 0)
        client = str(row.get("Client") or "").strip()
        ref = norm_ref(row.get("Ref"))

        values.append(
            "("
            + ", ".join(
                [
                    q(owner) + "::uuid",
                    q(client),
                    q(ref),
                    f"{total:.2f}",
                    f"{installment:.2f}",
                    str(paid_count),
                    str(total_inst),
                    q(start),
                    "7",
                    q(json.dumps(dates)) + "::jsonb",
                    q(json.dumps(notes)) + "::jsonb",
                    "'[]'::jsonb",
                    "0",
                    q("Imported status: Open"),
                    q(provider_type),
                    q(provider_name) if provider_name else "NULL",
                    f"{factoring:.2f}",
                    "false",
                ]
            )
            + ")"
        )

    lines.append(",\n".join(values) + ";")
    lines += [
        "",
        "COMMIT;",
        "",
        f"-- open={len(open_import)} excel_remaining≈{excel_remaining:.2f} "
        f"closed_lexi={len(closed[owner_lexi])} closed_neno={len(closed[owner_neno])}",
    ]

    sql_text = "\n".join(lines)
    OUT_SQL.write_text(sql_text, encoding="utf-8")
    print(f"Wrote {OUT_SQL}")
    print(f"Open inserts={len(open_import)} excel_remaining={excel_remaining:.2f}")
    print(
        f"Closed deletes Lexi={len(closed[owner_lexi])} Neno={len(closed[owner_neno])}"
    )
    for owner, row in open_import:
        tag = "LEXI" if owner == owner_lexi else "NENO"
        print(
            f"  {tag} {norm_ref(row.get('Ref'))} {row.get('Client')} rem={row.get('Remaining')}"
        )

    actionable = "\n".join(ln for ln in lines if not ln.strip().startswith("--"))
    if owner_petar in actionable:
        raise SystemExit("ABORT: Petar owner_id in actionable SQL")

    before = db_query(
        "SELECT count(*)::int AS cnt, coalesce(sum(id),0)::bigint AS id_sum "
        f"FROM public.loans WHERE owner_id = '{owner_petar}'"
    )
    print("Petar before:", before)

    proc = subprocess.run(
        ["npx", "supabase", "db", "query", "--linked", sql_text],
        cwd=str(REPO),
        capture_output=True,
        text=True,
        shell=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"Apply failed:\n{proc.stderr}\n{proc.stdout}")
    print("SQL applied OK")

    after = db_query(
        "SELECT count(*)::int AS cnt, coalesce(sum(id),0)::bigint AS id_sum "
        f"FROM public.loans WHERE owner_id = '{owner_petar}'"
    )
    print("Petar after:", after)
    if before != after:
        raise SystemExit("ABORT: Petar loans changed")

    summary = db_query(
        """
        SELECT owner_id::text AS owner_id,
               count(*)::int AS loans,
               count(*) FILTER (WHERE paid_count < total_installments)::int AS open_loans,
               round(sum(
                 CASE WHEN paid_count < total_installments
                   THEN GREATEST(0, (total_installments - paid_count)::numeric * installment
                                   - COALESCE(partial_paid_amount,0))
                   ELSE 0 END
               )::numeric, 2) AS open_remaining
        FROM public.loans
        WHERE coalesce(hidden,false)=false
        GROUP BY owner_id
        ORDER BY owner_id
        """
    )
    total_rem = 0.0
    print("Summary:")
    for row in summary:
        print(row)
        total_rem += float(row["open_remaining"] or 0)
    print(f"Total open remaining ≈ {total_rem:.2f} (target ~981416.25)")


if __name__ == "__main__":
    main()
